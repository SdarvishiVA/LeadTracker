"""
VA Capital — Lead Tracker Aggregator

Reads each broker's "Lead Log" tab directly from Box, combines them,
computes summary stats, and writes site/dashboard_data.json for the
website to display.

AUTH MODE
---------
This script needs three GitHub Actions secrets to talk to Box for real:
    BOX_CLIENT_ID
    BOX_CLIENT_SECRET
    BOX_ENTERPRISE_ID
These come from the Box Developer Console app (Client Credentials Grant)
once it's created and authorized by your Box admin. Until those secrets
exist, this script runs in DEMO MODE: it uses the two local sample files
in scripts/sample_data/ instead of calling Box, so the site has something
real to show while the Box admin step is pending.

Once the three secrets are set in the repo (Settings > Secrets and
variables > Actions), this script automatically switches to live mode —
no code change needed.
"""
import os
import io
import json
import datetime
from collections import defaultdict

import requests
import openpyxl

# ---------------------------------------------------------------------------
# BROKER REGISTRY
# Add one entry per broker here. `box_file_id` is the file's Box ID (visible
# in its Box URL: https://app.box.com/file/<this number>). Get it via the
# Box web app "Copy Box File Link" button, or Box:search_files_keyword.
# ---------------------------------------------------------------------------
BROKERS = [
    {"broker": "Suzana", "box_file_id": "2412094141422"},
    {"broker": "Raph",   "box_file_id": "2415423203380"},
    {"broker": "Kevin", "box_file_id": "2426484861293"},
    # {"broker": "NextBroker", "box_file_id": "PASTE_ID_HERE"},
]

REASONS = ["Insufficient Funds/Equity", "Not Serious/Just Browsing",
           "Residential Only (Not Multi-Res/Commercial)", "Wrong Deal Size or Type",
           "Timing Not Ready", "Unreachable", "Other"]
SOURCES_LIST = ["Instagram Ad", "COI Referral", "Referral", "Property Listing",
                "Cold Outreach", "LinkedIn", "Networking Event", "Past Client/Contact", "Other"]

DEMO_MODE = not all(os.environ.get(k) for k in
                     ("BOX_CLIENT_ID", "BOX_CLIENT_SECRET", "BOX_ENTERPRISE_ID"))


# ---------------------------------------------------------------------------
# Box auth + download (live mode only)
# ---------------------------------------------------------------------------
def get_box_token():
    resp = requests.post(
        "https://api.box.com/oauth2/token",
        data={
            "client_id": os.environ["BOX_CLIENT_ID"],
            "client_secret": os.environ["BOX_CLIENT_SECRET"],
            "grant_type": "client_credentials",
            "box_subject_type": "enterprise",
            "box_subject_id": os.environ["BOX_ENTERPRISE_ID"],
        },
        timeout=30,
    )
    resp.raise_for_status()
    return resp.json()["access_token"]


def download_box_file(file_id, token):
    resp = requests.get(
        f"https://api.box.com/2.0/files/{file_id}/content",
        headers={"Authorization": f"Bearer {token}"},
        timeout=60,
    )
    resp.raise_for_status()
    return io.BytesIO(resp.content)


# ---------------------------------------------------------------------------
# Reading a broker's Lead Log
# ---------------------------------------------------------------------------
def read_broker(broker, file_like_or_path):
    try:
        wb = openpyxl.load_workbook(file_like_or_path, data_only=True)
        if "Lead Log" not in wb.sheetnames:
            return [], f"'{broker}': no 'Lead Log' tab found"
        ws = wb["Lead Log"]
        headers = [c.value for c in ws[1]]
        required = ["Date", "Lead Name", "Source of Lead", "Responded?", "Qualified?", "Reason Not Qualified"]
        missing = [h for h in required if h not in headers]
        if missing:
            return [], f"'{broker}': missing expected column(s) {missing}"
        idx = {h: headers.index(h) for h in headers if h}
        rows = []
        for r in range(2, ws.max_row + 1):
            name = ws.cell(row=r, column=idx["Lead Name"] + 1).value
            if not name:
                continue
            date = ws.cell(row=r, column=idx["Date"] + 1).value
            rows.append({
                "broker": broker,
                "date": date.strftime("%Y-%m-%d") if isinstance(date, datetime.datetime) else None,
                "_date_obj": date if isinstance(date, datetime.datetime) else None,
                "name": str(name),
                "company": ws.cell(row=r, column=idx["Company"] + 1).value if "Company" in idx else None,
                "prospect_type": ws.cell(row=r, column=idx["Prospect Type"] + 1).value if "Prospect Type" in idx else None,
                "type": ws.cell(row=r, column=idx["Type"] + 1).value if "Type" in idx else None,
                "owns_commercial": ws.cell(row=r, column=idx["invested before?"] + 1).value if "invested before?" in idx else None,
                "source": ws.cell(row=r, column=idx["Source of Lead"] + 1).value,
                "responded": ws.cell(row=r, column=idx["Responded?"] + 1).value,
                "qualified": ws.cell(row=r, column=idx["Qualified?"] + 1).value,
                "reason": ws.cell(row=r, column=idx["Reason Not Qualified"] + 1).value,
            })
        return rows, None
    except Exception as e:
        return [], f"'{broker}': failed to read — {type(e).__name__}: {e}"


def week_of(d):
    monday = d - datetime.timedelta(days=d.weekday())
    return f'{monday.strftime("%b %-d")} - {(monday + datetime.timedelta(days=5)).strftime("%b %-d")}', monday


# ---------------------------------------------------------------------------
# Build the aggregate JSON
# ---------------------------------------------------------------------------
def build():
    all_rows = []
    errors = []
    ok_brokers = []

    token = None if DEMO_MODE else get_box_token()

    for b in BROKERS:
        broker = b["broker"]
        if DEMO_MODE:
            path = f"scripts/sample_data/{broker.lower()}.xlsx"
            source = path
        else:
            try:
                source = download_box_file(b["box_file_id"], token)
            except Exception as e:
                errors.append(f"'{broker}': Box download failed — {type(e).__name__}: {e}")
                continue

        rows, err = read_broker(broker, source)
        if err:
            errors.append(err)
            continue
        all_rows.extend(rows)
        ok_brokers.append(broker)

    total = len(all_rows)
    responded = sum(1 for r in all_rows if r["responded"] == "Yes")
    qualified = sum(1 for r in all_rows if r["qualified"] == "Yes")

    by_broker = []
    for b in BROKERS:
        name = b["broker"]
        if name not in ok_brokers:
            continue
        b_rows = [r for r in all_rows if r["broker"] == name]
        n = len(b_rows)
        resp = sum(1 for r in b_rows if r["responded"] == "Yes")
        qual = sum(1 for r in b_rows if r["qualified"] == "Yes")
        by_broker.append({
            "broker": name, "leads": n, "responded": resp,
            "response_rate": round(resp / n, 4) if n else 0,
            "qualified": qual,
            "qualify_rate": round(qual / n, 4) if n else 0,
        })

    by_source = []
    for s in SOURCES_LIST:
        s_rows = [r for r in all_rows if r["source"] == s]
        n = len(s_rows)
        if n == 0:
            continue
        resp = sum(1 for r in s_rows if r["responded"] == "Yes")
        by_source.append({"source": s, "leads": n, "responded": resp,
                           "response_rate": round(resp / n, 4) if n else 0})

    weeks = defaultdict(lambda: defaultdict(int))
    week_order = []
    for r in all_rows:
        if not r["_date_obj"]:
            continue
        label, monday = week_of(r["_date_obj"])
        if label not in week_order:
            week_order.append((monday, label))
        weeks[label]["total"] += 1
        weeks[label][r["broker"]] += 1
        if r["responded"] == "Yes":
            weeks[label]["responded"] += 1
    week_order.sort()
    by_week = []
    for _, label in week_order:
        row = {"week_of": label, "total": weeks[label]["total"], "responded": weeks[label]["responded"]}
        for b in BROKERS:
            row[b["broker"]] = weeks[label].get(b["broker"], 0)
        by_week.append(row)

    by_reason = [{"reason": r, "count": sum(1 for x in all_rows if x["reason"] == r)} for r in REASONS]

    leads_out = sorted(
        [
            {
                "broker": r["broker"],
                "date": r["date"],
                "name": r["name"],
                "company": r["company"],
                "prospect_type": r["prospect_type"],
                "type": r["type"],
                "owns_commercial": r["owns_commercial"],
                "source": r["source"],
                "responded": r["responded"],
                "qualified": r["qualified"],
                "reason": r["reason"],
            }
            for r in all_rows
        ],
        key=lambda x: x["date"] or "",
        reverse=True,
    )

    return {
        "generated_at": datetime.datetime.now(datetime.timezone.utc).isoformat(),
        "demo_mode": DEMO_MODE,
        "sources_read_ok": ok_brokers,
        "errors": errors,
        "totals": {
            "leads": total, "responded": responded,
            "response_rate": round(responded / total, 4) if total else 0,
            "qualified": qualified,
            "qualify_rate": round(qualified / total, 4) if total else 0,
        },
        "by_broker": by_broker,
        "by_source": by_source,
        "by_week": by_week,
        "by_disqualification_reason": by_reason,
        "leads": leads_out,
    }


if __name__ == "__main__":
    output = build()
    out_path = os.path.join(os.path.dirname(__file__), "..", "dashboard_data.json")
    with open(out_path, "w") as f:
        json.dump(output, f, indent=2)
    print(f"Wrote {out_path}")
    print(f"DEMO_MODE = {DEMO_MODE}")
    print(f"sources_read_ok = {output['sources_read_ok']}")
    print(f"errors = {output['errors'] or 'none'}")
    print(f"totals = {output['totals']}")
